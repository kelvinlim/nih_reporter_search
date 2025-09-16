#!/usr/bin/env python3
"""
NIH Reporter Search Tool

This script searches the NIH Reporter API for funding information
for a list of names provided in a YAML file.
"""

import json
import yaml
import requests
import argparse
import csv
import os
from datetime import datetime
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict


class NIHReporterSearcher:
    """Class to handle NIH Reporter API searches."""
    
    def __init__(self):
        self.base_url = "https://api.reporter.nih.gov/v2/projects/search"
        self.session = requests.Session()
        self.session.headers.update({
            'Content-Type': 'application/json',
            'User-Agent': 'NIH-Reporter-Search-Tool/1.0'
        })
    
    def search_person(self, name: str, organization: str = None) -> List[Dict[str, Any]]:
        """
        Search for funding information for a specific person.
        
        Args:
            name: The name of the person to search for
            organization: Optional organization to filter by
            
        Returns:
            List of project dictionaries from the API
        """
        # Prepare the search criteria
        criteria = {
            "pi_names": [{"any_name": name}]
        }
        
        # Add organization filter if provided
        if organization:
            criteria["org_names"] = [organization]
        
        search_criteria = {
            "criteria": criteria,
            "offset": 0,
            "limit": 500,  # Maximum allowed by API
            "sort_field": "project_start_date",
            "sort_order": "desc"
        }
        
        try:
            response = self.session.post(self.base_url, json=search_criteria)
            response.raise_for_status()
            
            data = response.json()
            return data.get('results', [])
            
        except requests.exceptions.RequestException as e:
            print(f"Error searching for {name}: {e}")
            return []
        except json.JSONDecodeError as e:
            print(f"Error parsing response for {name}: {e}")
            return []
    
    def process_funding_data(self, projects: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Process funding data and organize projects directly.
        
        Args:
            projects: List of project dictionaries from API
            
        Returns:
            Dictionary with project data and totals
        """
        total_direct_costs = 0.0
        total_costs = 0.0
        processed_projects = []
        
        for project in projects:
            # Extract project information
            project_id = project.get('project_num', 'Unknown')
            title = project.get('project_title', 'Unknown Title')
            # Use budget dates instead of project dates
            start_date = project.get('budget_start', '') or project.get('project_start_date', '')
            end_date = project.get('budget_end', '') or project.get('project_end_date', '')
            
            # Extract funding information (handle None values and different field names)
            direct_costs = project.get('direct_cost_amt', 0.0) or 0.0
            indirect_costs = project.get('indirect_cost_amt', 0.0) or 0.0
            award_amount = project.get('award_amount', 0.0) or 0.0
            
            # Calculate total costs: if award_amount is available, use it; otherwise sum direct + indirect
            if award_amount > 0:
                project_total_costs = award_amount
            else:
                project_total_costs = direct_costs + indirect_costs
            
            # Add to totals
            total_direct_costs += direct_costs
            total_costs += project_total_costs
            
            # Add project to list
            processed_projects.append({
                'project_id': project_id,
                'title': title,
                'start_date': start_date,  # Budget start date (or project start date if budget not available)
                'end_date': end_date,      # Budget end date (or project end date if budget not available)
                'budget_start_date': project.get('budget_start', ''),
                'budget_end_date': project.get('budget_end', ''),
                'project_start_date': project.get('project_start_date', ''),
                'project_end_date': project.get('project_end_date', ''),
                'direct_costs': direct_costs,
                'indirect_costs': indirect_costs,
                'award_amount': award_amount,
                'total_costs': project_total_costs
            })
        
        return {
            'total_direct_costs': total_direct_costs,
            'total_costs': total_costs,
            'project_count': len(processed_projects),
            'projects': processed_projects
        }
    
    def create_summary_csv(self, results: Dict[str, Any], output_file: str = 'summary.csv'):
        """
        Create a summary CSV file with key metrics for each person.
        
        Args:
            results: Dictionary containing search results
            output_file: Path to output CSV file
        """
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Name', 'Total_Direct_Costs', 'Total_Costs', 'Most_Recent_Year', 'Total_Projects', 
                         'Current_Direct_Costs', 'Current_Total_Costs', 'Current_Projects']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            # Sort results by last name for CSV output
            sorted_results = self._sort_results_by_last_name(results)
            
            for name, data in sorted_results.items():
                # Format name as "Last Name, First Name" for CSV
                formatted_name = self._format_name_last_first(name)
                
                # Get totals directly from the data structure
                total_direct = data.get('total_direct_costs', 0.0)
                total_costs = data.get('total_costs', 0.0)
                total_projects = data.get('total_projects', 0)
                projects = data.get('projects', [])
                
                # Calculate current funding (projects active today)
                current_direct = 0.0
                current_total = 0.0
                current_project_count = 0
                today = datetime.now().date()
                
                for project in projects:
                    start_date = project.get('start_date', '')
                    end_date = project.get('end_date', '')
                    
                    if start_date and end_date:
                        try:
                            # Parse start and end dates
                            project_start = datetime.strptime(start_date, '%Y-%m-%dT%H:%M:%S').date()
                        except ValueError:
                            try:
                                project_start = datetime.strptime(start_date, '%Y-%m-%d').date()
                            except ValueError:
                                continue
                        
                        try:
                            project_end = datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%S').date()
                        except ValueError:
                            try:
                                project_end = datetime.strptime(end_date, '%Y-%m-%d').date()
                            except ValueError:
                                continue
                        
                        # Check if project is currently active
                        if project_start <= today <= project_end:
                            current_direct += project.get('direct_costs', 0.0)
                            current_total += project.get('total_costs', 0.0)
                            current_project_count += 1
                
                # Find most recent year based on project end dates
                most_recent_year = None
                if projects:
                    # Look through all projects to find the most recent end date
                    latest_end_date = None
                    for project in projects:
                        end_date = project.get('end_date', '')
                        if end_date:
                            try:
                                # Parse the end date and compare
                                project_end_date = datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%S')
                                if latest_end_date is None or project_end_date > latest_end_date:
                                    latest_end_date = project_end_date
                            except ValueError:
                                # Try alternative date format
                                try:
                                    project_end_date = datetime.strptime(end_date, '%Y-%m-%d')
                                    if latest_end_date is None or project_end_date > latest_end_date:
                                        latest_end_date = project_end_date
                                except ValueError:
                                    continue
                    
                    if latest_end_date:
                        most_recent_year = latest_end_date.year
                
                writer.writerow({
                    'Name': formatted_name,
                    'Total_Direct_Costs': f"{total_direct:,.2f}",
                    'Total_Costs': f"{total_costs:,.2f}",
                    'Most_Recent_Year': most_recent_year or 'N/A',
                    'Total_Projects': total_projects,
                    'Current_Direct_Costs': f"{current_direct:,.2f}",
                    'Current_Total_Costs': f"{current_total:,.2f}",
                    'Current_Projects': current_project_count
                })
        
        print(f"Summary saved to: {output_file}")
    
    def create_summary_excel(self, results: Dict[str, Any], output_file: str) -> None:
        """
        Create an Excel summary file with funding data.
        
        Args:
            results: Dictionary with results for each person
            output_file: Path to output Excel file
        """
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "NIH Funding Summary"
        
        # Define headers
        headers = ['Name', 'Total_Direct_Costs', 'Total_Costs', 'Most_Recent_Year', 
                  'Total_Projects', 'Current_Direct_Costs', 'Current_Total_Costs', 'Current_Projects']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Sort results by last name for Excel output
        sorted_results = self._sort_results_by_last_name(results)
        
        # Write data rows
        row = 2
        for name, data in sorted_results.items():
            # Format name as "Last Name, First Name" for Excel
            formatted_name = self._format_name_last_first(name)
            
            # Get totals directly from the data structure
            total_direct = data.get('total_direct_costs', 0.0)
            total_costs = data.get('total_costs', 0.0)
            total_projects = data.get('total_projects', 0)
            projects = data.get('projects', [])
            
            # Calculate current funding (projects active today)
            current_direct = 0.0
            current_total = 0.0
            current_project_count = 0
            today = datetime.now().date()
            
            for project in projects:
                start_date = project.get('start_date', '')
                end_date = project.get('end_date', '')
                
                if start_date and end_date:
                    try:
                        # Parse start and end dates
                        project_start = datetime.strptime(start_date, '%Y-%m-%dT%H:%M:%S').date()
                    except ValueError:
                        try:
                            project_start = datetime.strptime(start_date, '%Y-%m-%d').date()
                        except ValueError:
                            continue
                    
                    try:
                        project_end = datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%S').date()
                    except ValueError:
                        try:
                            project_end = datetime.strptime(end_date, '%Y-%m-%d').date()
                        except ValueError:
                            continue
                    
                    # Check if project is currently active
                    if project_start <= today <= project_end:
                        current_direct += project.get('direct_costs', 0.0)
                        current_total += project.get('total_costs', 0.0)
                        current_project_count += 1
            
            # Find most recent year based on project end dates
            most_recent_year = None
            if projects:
                # Look through all projects to find the most recent end date
                latest_end_date = None
                for project in projects:
                    end_date = project.get('end_date', '')
                    if end_date:
                        try:
                            # Parse the end date and compare
                            project_end_date = datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%S')
                            if latest_end_date is None or project_end_date > latest_end_date:
                                latest_end_date = project_end_date
                        except ValueError:
                            # Try alternative date format
                            try:
                                project_end_date = datetime.strptime(end_date, '%Y-%m-%d')
                                if latest_end_date is None or project_end_date > latest_end_date:
                                    latest_end_date = project_end_date
                            except ValueError:
                                continue
                
                if latest_end_date:
                    most_recent_year = latest_end_date.year
            
            # Write row data
            ws.cell(row=row, column=1, value=formatted_name)
            ws.cell(row=row, column=2, value=total_direct)
            ws.cell(row=row, column=3, value=total_costs)
            ws.cell(row=row, column=4, value=most_recent_year or 'N/A')
            ws.cell(row=row, column=5, value=total_projects)
            ws.cell(row=row, column=6, value=current_direct)
            ws.cell(row=row, column=7, value=current_total)
            ws.cell(row=row, column=8, value=current_project_count)
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format number columns
        for row in range(2, row):
            # Format currency columns (2, 3, 6, 7)
            for col in [2, 3, 6, 7]:
                cell = ws.cell(row=row, column=col)
                if cell.value and cell.value != 0:
                    cell.number_format = '#,##0.00'
        
        # Save the workbook
        wb.save(output_file)
        print(f"Excel summary saved to: {output_file}")
    
    def _sort_results_by_last_name(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """
        Sort results alphabetically by last name.
        
        Args:
            results: Dictionary with results for each person
            
        Returns:
            Dictionary sorted by last name
        """
        def get_last_name(name: str) -> str:
            """Extract last name from full name."""
            # Split by space and take the last part as last name
            parts = name.strip().split()
            if not parts:
                return ""
            return parts[-1]
        
        # Sort the results by last name
        sorted_items = sorted(results.items(), key=lambda x: get_last_name(x[0]))
        return dict(sorted_items)
    
    def _format_name_last_first(self, name: str) -> str:
        """
        Format name as "Last Name, First Name".
        
        Args:
            name: Full name in "First Name Last Name" format
            
        Returns:
            Name formatted as "Last Name, First Name"
        """
        parts = name.strip().split()
        if len(parts) < 2:
            return name  # Return as-is if can't split properly
        
        # Last name is the last part, everything else is first name
        last_name = parts[-1]
        first_name = ' '.join(parts[:-1])
        
        return f"{last_name}, {first_name}"
    
    def search_names_from_yaml(self, yaml_file: str, extra_text: str = "") -> Dict[str, Any]:
        """
        Search for funding information for all names in a YAML file.
        
        Args:
            yaml_file: Path to YAML file containing names
            extra_text: Optional organization name to filter by (e.g., "University of Minnesota")
            
        Returns:
            Dictionary with results for each person
        """
        try:
            with open(yaml_file, 'r') as f:
                data = yaml.safe_load(f)
            
            names = data.get('names', [])
            if not names:
                print("No names found in YAML file")
                return {}
            
            results = {}
            
            for name in names:
                # Clean up the name (remove periods from middle initials for better API matching)
                clean_name = name.replace('.', '').replace('  ', ' ')
                print(f"Searching for: {clean_name}" + (f" at {extra_text}" if extra_text else ""))
                projects = self.search_person(clean_name, extra_text)
                processed_data = self.process_funding_data(projects)
                
                results[name] = {
                    'total_projects': processed_data['project_count'],
                    'total_direct_costs': processed_data['total_direct_costs'],
                    'total_costs': processed_data['total_costs'],
                    'projects': processed_data['projects'],
                    'search_timestamp': datetime.now().isoformat(),
                    'search_name_used': clean_name,
                    'organization_filter': extra_text if extra_text else None
                }
                
                print(f"Found {len(projects)} projects for {name}" + (f" at {extra_text}" if extra_text else ""))
            
            # Sort results alphabetically by last name
            return self._sort_results_by_last_name(results)
            
        except FileNotFoundError:
            print(f"Error: YAML file '{yaml_file}' not found")
            return {}
        except yaml.YAMLError as e:
            print(f"Error parsing YAML file: {e}")
            return {}


def main():
    """Main function to run the NIH Reporter search."""
    parser = argparse.ArgumentParser(description='Search NIH Reporter API for funding information')
    parser.add_argument('yaml_file', help='Path to YAML file containing names to search')
    parser.add_argument('-o', '--output', help='Output JSON file (default: based on YAML filename)')
    parser.add_argument('--extra', help='Organization name to filter by (e.g., "University of Minnesota")')
    
    args = parser.parse_args()
    
    # Generate output filename based on YAML file basename
    if args.output:
        output_file = args.output
    else:
        yaml_basename = os.path.splitext(os.path.basename(args.yaml_file))[0]
        output_file = f"{yaml_basename}_results.json"
    
    # Create searcher instance
    searcher = NIHReporterSearcher()
    
    # Search for names
    extra_text = args.extra or ""
    results = searcher.search_names_from_yaml(args.yaml_file, extra_text)
    
    if results:
        # Save results to JSON file
        with open(output_file, 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"\nResults saved to: {output_file}")
        
        # Create summary CSV
        summary_file = output_file.replace('.json', '_summary.csv')
        searcher.create_summary_csv(results, summary_file)
        
        # Create summary Excel
        excel_file = output_file.replace('.json', '_summary.xlsx')
        searcher.create_summary_excel(results, excel_file)
        
        # Print summary
        print("\nSummary:")
        for name, data in results.items():
            total_direct = data.get('total_direct_costs', 0.0)
            total_costs = data.get('total_costs', 0.0)
            print(f"{name}: {data['total_projects']} projects, "
                  f"${total_direct:,.2f} direct costs, "
                  f"${total_costs:,.2f} total costs")
    else:
        print("No results found")


if __name__ == "__main__":
    main()
